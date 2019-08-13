import threading
from queue import Queue

from django.core.paginator import Paginator
from django.shortcuts import render
from django.http.response import HttpResponse, JsonResponse
from django.contrib.auth.models import Group, User
from django.views.generic import DetailView
from .models import Lawyer, Case, Status, AssignLinks, SeenDetails
from django.db.models import Count
from django.core import management
from lxml.html import fromstring
from lxml import etree
from django.core.files.storage import FileSystemStorage
from openpyxl import load_workbook
from multiprocessing.pool import ThreadPool
# from django.http import HttpResponse
from datetime import datetime, timedelta
from django.utils import formats
import string
import json

from scrappers.second import get_texasbar_link
from scrappers.third import get_texasbar_details
from scrappers.forth import get_cases
from scrappers.fifth import get_case, update_case
from courts import settings
from django_rq import job
from django.contrib.auth.decorators import login_required
from django.views.generic import TemplateView
from django.contrib import auth
from django.shortcuts import render, redirect, get_object_or_404


@job('high', timeout=7200)
def update_all_data(excel_file_path):
    print('****************************>>>>>>>>>>>>2')

    def process_attorney_name(name):
        name = str(name)
        name = name.replace('  ', ' ')
        if len(name.split()) == 2:
            if name.split()[0].endswith('.') and name.split()[1].endswith('.'):
                return ''
            return name
        name = name.split(',')[0]  # A. Robert Lamb, Jr.
        if name.startswith('Hon.'):
            name = name[5:]
        if name.startswith('Mr.'):
            name = name[4:]
        if name.startswith('Mrs.'):
            name = name[5:]
        if name.startswith('Ms.'):
            name = name[4:]
        for _ in range(3):
            for initial in list(string.ascii_uppercase):
                if name.startswith(initial + '.'):
                    name = name[3:]

        if len(name.split()) < 2:
            return ''
        return name.split()[0] + ' ' + name.split()[-1]

    wb2 = load_workbook(settings.BASE_DIR + excel_file_path)
    sheet = wb2.active
    db_bar_cards = Lawyer.objects.values('bar_card')



    db_bar_cards = [bar_card['bar_card'] for bar_card in db_bar_cards]
    db_case_numbers = list(Case.objects.values('case_number', 'lawyer_id'))

    print('Read Attorneys from Excel Sheet')
    attorneys = []
    bar_card_numbers = []
    for row_id in range(2, sheet.max_row + 1):
        name = process_attorney_name(sheet['A' + str(row_id)].value)
        bar_card = sheet['B' + str(row_id)].value

        if name and not bar_card:
            attorneys.append(name)
        if bar_card:
            bar_card_numbers.append(bar_card)

    print('Get all Links of Attorneys without Bar Card Number in Excel Sheet')
    simples = []
    pool = ThreadPool(32)
    for _, simple in enumerate(pool.imap_unordered(get_texasbar_link, attorneys)):
        print(simple)
        simples.append(simple)

    print('Update new Attorneys')
    simples = [simple for simple in simples if 'Link' in simple]



    pool = ThreadPool(32)
    for _, row in enumerate(pool.imap_unordered(get_texasbar_details, simples)):
        bar_card_number = row['Bar Card']
        bar_card_numbers.append(bar_card_number)
        if bar_card_number not in db_bar_cards:
            obj = Lawyer()
            obj.bar_card = row.get('Bar Card', '')
            obj.first_name = row.get('First Name', '')
            obj.last_name = row.get('Last Name', '')
            obj.full_name = row.get('Full Name', '')
            obj.status = row.get('Status', '')
            obj.company = row.get('Company', '')
            obj.practice_areas = row.get('Practice Areas', '')
            obj.address = row.get('Address', '')
            obj.practice_location = row.get('Practice Location', '')
            obj.gmaps_img = row.get('Gmaps img', '')
            obj.profile_img = row.get('Profile img', '')
            obj.license_date = row.get('License Date', '')
            obj.statutory_profile_date = row.get('Statutory Profile Date', '')

            try:
                obj.save()
            except:
                pass
        print(row)

    print('Get New Cases')
    all_case = []
    pool = ThreadPool(32)
    totalNum = 0
    for r, chunk in enumerate(pool.imap_unordered(get_cases, bar_card_numbers)):
        try:
            bar_card, cases = chunk
            print(r, bar_card, len(cases))
            totalNum += len(cases)
            new_cases = [case for case in cases if case['Bar Card'] and case['Case Number']
                         and case['Case Number'] != 'Your search found no results. Try broadening your search criteria.'
                         and not db_case_numbers.__contains__({'case_number': case['Case Number'],
                                                               'lawyer_id': case['Bar Card']})]
            if new_cases:
                print(r, bar_card, len(new_cases))
                all_case += new_cases
        except Exception as e:
            print(e)

    print('**********************************************************')
    print(totalNum)
    print('**********************************************************')
    print('Get all data of new Cases:', len(all_case))

    pool = ThreadPool(32)
    for r, c in enumerate(pool.imap_unordered(get_case, all_case)):
        print(r, c['Bar Card'], c['Case Number'])
        try:
            a = Case(lawyer_id=c['Bar Card'], appellate_court=c['Appellate Court'],
                     coa_case_number=c['COA Case Number'], case_number=c['Case Number'],
                     case_type=c['Case Type'], date_filed=c['Date Filed'], style=c['Style'],
                     trial_court=c['Trial Court'], trial_court_case_number=c['Trial Court Case Number'],
                     trial_court_county=c['Trial Court County'], v=c['v.'], appellate_briefs=c['briefs'],
                     calendars=c['calendars'], case_events=c['events'], parties=c['parties'],
                     trial_court_information=c['trial'])
        except Exception as e:
            print(1, e)
            try:
                a = Case(lawyer_id=c['Bar Card'], appellate_court=c['Appellate Court'],
                         coa_case_number=c['COA Case Number'], case_number=c['Case Number'],
                         case_type=c['Case Type'], date_filed=c['Date Filed'], style=c['Style'],
                         trial_court=c['Trial Court'], trial_court_case_number=c['Trial Court Case Number'],
                         trial_court_county=c['Trial Court County'], v=c['v.'], appellate_briefs='',
                         calendars='', case_events='', parties='', trial_court_information='')
            except Exception as e:
                print(2, e)

        try:
            if a.case_type not in ['11.07', '11.07 HC', '11.071', '1107-HC']:
                a.case_type = ' '.join([word[0].upper() + word[1:].lower() for word in a.case_type.split() if
                                        word not in ['to', 'or', 'for', 'of', 'with', 'under']])
            a.save()
        except Exception as e:
            print(3, e)



    lawyers_list = Lawyer.objects.annotate(cases_count=Count('case')).order_by('-cases_count')
    file = open(settings.MEDIA_ROOT + '/lawyers.txt', 'w', encoding='utf-8')
    for lawyer in lawyers_list:
        file.write(','.join([lawyer.bar_card, lawyer.full_name, str(lawyer.cases_count)]) + '\n')
    file.close()

    cases_list = Case.objects.values('case_type').exclude(case_type='').order_by('case_type')
    file = open(settings.MEDIA_ROOT + '/case_types.txt', 'w', encoding='utf-8')
    file.write(str(','.join(sorted(set(list([case['case_type'] for case in cases_list]))))))
    file.close()

    file = open(settings.MEDIA_ROOT + '/case_types.txt', encoding='utf-8', errors='ignore')
    cases = file.read().split(',')
    file.close()

    for case in cases:
        lawyers_list = Lawyer.objects.filter(case__case_type=case).annotate(cases_count=Count('case')).order_by(
            '-cases_count')
        file_name = case.replace('\\', ' ').replace('/', ' ').replace('&', '_').replace('<', ' ').replace('>', ' ')
        file_name = file_name.replace(':', ' ').replace('"', ' ').replace('|', ' ').replace('?', ' ').replace('*', ' ')
        file = open(settings.MEDIA_ROOT + '/' + file_name + '.txt', 'w', encoding='utf-8')
        for lawyer in lawyers_list:
            a = dict()
            a['bar_card'] = lawyer.bar_card
            a['full_name'] = lawyer.full_name
            a['cases_count'] = lawyer.cases_count
            a['all_cases_count'] = lawyer.case_set.count()
            file.write(json.dumps(a) + '\n')
        file.close()
    print('update_all_data() finished')
    x = q.get()


@job
def update_database_cases():
    try:
        current, created = Status.objects.get_or_create(id=1)

        pool = ThreadPool(32)
        if current.current_update > Case.objects.all().count():
            current.current_update = 0
        db_cases = Case.objects.all()
        db_cases = db_cases[current.current_update:current.current_update + 3000]

        for r, chunk in enumerate(pool.imap_unordered(update_case, db_cases)):
            print(r, chunk)
        current.current_update = current.current_update + 3000
        current.save()

    except Exception as e:
        print(e)


def home_page(request):
    if not request.session.session_key:
        request.session['paid'] = "false"
    if not request.user.username:
        context = {'user_name': 'Not logined User', "session": request.session['paid'], 'groupname': 'no group'}
        request.session['paid'] = "false"
    else:

        user_date = request.user.date_joined + timedelta(days=30)
        user_date = formats.date_format(user_date, "SHORT_DATETIME_FORMAT")
        now = datetime.now()
        now = formats.date_format(now, "SHORT_DATETIME_FORMAT")

        expiredate = datetime.strptime('Feb 1 2019  11:13PM', '%b %d %Y %I:%M%p')
        if now > user_date:
            boolean = "true"
            disAllow_groups = Group.objects.get(name='Disallow')
            disAllow_groups.user_set.add(request.user)
            Addgroup = Group.objects.get(name='Allowed')
            Addgroup.user_set.remove(request.user)
        else:
            boolean = "false"
            disAllow_groups = Group.objects.get(name='Allowed')
            disAllow_groups.user_set.add(request.user)
            Addgroup = Group.objects.get(name='Disallow')
            Addgroup.user_set.remove(request.user)
        #     g = Group.objects.get(name='Disallow')
        #     g.user_set.add(request.user)
        groupname = []
        for g in request.user.groups.all():
            groupname.append(g.name)
        allow = "Allowed"
        if allow in groupname:
            request.session['paid'] = "true"
            context = {'user_name': request.user.username, "session": request.session['paid'], 'groupname': groupname,
                       'expiredate': user_date, 'current': now}
        else:
            request.session['paid'] = "false"
            if request.user.username == "admin":
                request.session['paid'] = "true"
            context = {'user_name': request.user.username, "session": request.session['paid'], 'groupname': groupname,
                       'expiredate': user_date, 'current': now}

    return render(request, 'home.html', context=context)


q = Queue()


def add_lawyer(request):
    if request.method == 'POST':
        if 'attorneys_name' in request.FILES:
            attorney_file = request.FILES['attorneys_name']
            fs = FileSystemStorage()
            filename = fs.save(attorney_file.name, attorney_file)
            uploaded_file_url = fs.url(filename)
            # result = update_all_data(uploaded_file_url)

            if q.qsize() < 1:
                print('****************************>>>>>>>>>>>>1')
                q.put('OK')
                update_thread = threading.Thread(
                    target=update_all_data, args=(uploaded_file_url,))
                update_thread.daemon = True
                update_thread.start()

            # import django_rq
            # queue = django_rq.get_queue('default')
            # queue.enqueue(
            #     update_all_data,
            #     args=(uploaded_file_url, ),
            #
            # )

            context = {'bar_car': 'success', 'update_type': 'all'}
        else:
            texasbar_link = request.POST['link']
            bar_card = management.call_command('add_lawyer_cases', texasbar_link)
            context = {'bar_car': bar_card, 'update_type': 'single'}
    else:
        context = {'bar_car': 'GET', 'update_type': ''}

    return render(request, 'add.html', context=context)


# Home Page View
def search_lawyers(request):
    user_date = datetime.now()
    user_date = formats.date_format(user_date, "SHORT_DATETIME_FORMAT")
    first_name = request.GET.get('first_name', '')
    last_name = request.GET.get('last_name', '')
    bar_card = request.GET.get('bar_card', '')
    request.session['casetype'] = " "
    if first_name or last_name or bar_card:
        lawyers_list = Lawyer.objects.filter(
            first_name__icontains=first_name).filter(
            last_name__icontains=last_name).filter(
            bar_card__icontains=bar_card).annotate(
            cases_count=Count('case')).order_by('-cases_count')
    else:
        try:
            # request.session['paid'] ="True"
            file = open(settings.MEDIA_ROOT + '/lawyers.txt', encoding='utf-8', errors='ignore')
            # lawyers_list = [{'bar_card': lawyer.split(',')[0],
            #                  'full_name': lawyer.split(',')[1],
            #                  'cases_count': lawyer.split(',')[2],
            #                  'first_name' : lawyer.split(',')[1].split(' ')[0]+" "+lawyer.split(',')[1].split(' ')[1],
            #                  'last_name'  :  lawyer.split(',')[1].replace(lawyer.split(',')[1].split(' ')[0]+" "+lawyer.split(',')[1].split(' ')[1],""),
            #                  } for lawyer in file.readlines()]
            lawyers_list = []
            index = 0
            for lawyer in file.readlines():
                index = index + 1
                list = {'index': index,
                        'bar_card': lawyer.split(',')[0],
                        'full_name': lawyer.split(',')[1],
                        'cases_count': lawyer.split(',')[2],
                        # 'first_name' : lawyer.split(',')[1].split()[0]+" "+lawyer.split(',')[1].split()[1],
                        #  'last_name'  :  lawyer.split(',')[1].replace(lawyer.split(',')[1].split(' ')[0]+" "+lawyer.split(',')[1].split(' ')[1],""),
                        }
                lawyers_list.append(list)
                # for i in range(5):
            #     lawyers_list.append([{'bar_card':'1','full_name':'1','cases_count':'end_page','first_name':'ddd','last_name':'sssss'}])
            # with open(settings.MEDIA_ROOT + '/lawyers.txt', mode="r",encoding="utf-8") as f:
            #     reader=csv.reader(f,delimiter=",")
            # for row_index, row_data in enumerate(reader):
            #     bar_card,full_name,end_page=row_data
            #     lawyers_list.append([{'bar_card':bar_card,'full_name':full_name,'cases_count':end_page,'first_name':full_name.split(' ')[0]+" "+full_name.split(' ')[1],'last_name':full_name.replace(full_name.split(' ')[0]+" "+full_name.split(' ')[1],"")}])
            # for lawyer in file.readlines():
            #     lawyers_list.append({'bar_card': lawyer.split(',')[0],'full_name': lawyer.split(',')[1],'cases_count': lawyer.split(',')[2],'first_name' : lawyer.split(',')[1].split(' ')[0]+" "+lawyer.split(',')[1].split(' ')[1],'last_name'  :  lawyer.split(',')[1].replace(lawyer.split(',')[1].split(' ')[0]+" "+lawyer.split(',')[1].split(' ')[1],""),'session': request.session['paid']})
            file.close()
        except:
            lawyers_list = []
    page = request.GET.get('page')
    paginator = Paginator(lawyers_list, 10)
    lawyers_list = paginator.get_page(page)
    page_number = lawyers_list.number
    request.session['pagenumber'] = page_number

    if not request.session.session_key:
        request.session['paid'] = "false"
    if not request.user.username:
        request.session['paid'] = "false"
    else:
        user_date = request.user.date_joined + timedelta(days=30)
        user_date = formats.date_format(user_date, "SHORT_DATETIME_FORMAT")
        now = datetime.now()
        now = formats.date_format(now, "SHORT_DATETIME_FORMAT")

        expiredate = datetime.strptime('Feb 1 2019  11:13PM', '%b %d %Y %I:%M%p')
        if now > user_date:
            boolean = "true"
            disAllow_groups = Group.objects.get(name='Disallow')
            disAllow_groups.user_set.add(request.user)
            Addgroup = Group.objects.get(name='Allowed')
            Addgroup.user_set.remove(request.user)
        else:
            boolean = "false"
            disAllow_groups = Group.objects.get(name='Allowed')
            disAllow_groups.user_set.add(request.user)
            Addgroup = Group.objects.get(name='Disallow')
            Addgroup.user_set.remove(request.user)
        groupname = []
        for g in request.user.groups.all():
            groupname.append(g.name)
        allow = "Allowed"
        if allow in groupname:
            request.session['paid'] = "true"
        else:
            request.session['paid'] = "false"
            if request.user.username == "admin":
                request.session['paid'] = "true"

    context = {'lawyers': lawyers_list, 'session': request.session['paid'], 'expiredate': user_date}
    return render(request, 'search_lawyers.html', context)


def lawyers_by_case(request):
    user_date = datetime.now()
    user_date = formats.date_format(user_date, "SHORT_DATETIME_FORMAT")
    case_type = request.GET.get('case_type', '')
    group = request.GET.get('group', '')

    if not request.session.session_key:
        request.session['paid'] = "false"
    if not request.user.username:
        request.session['paid'] = "false"
    else:
        user_date = request.user.date_joined + timedelta(days=30)
        user_date = formats.date_format(user_date, "SHORT_DATETIME_FORMAT")
        now = datetime.now()
        now = formats.date_format(now, "SHORT_DATETIME_FORMAT")

        expiredate = datetime.strptime('Feb 1 2019  11:13PM', '%b %d %Y %I:%M%p')
        if now > user_date:
            boolean = "true"
            disAllow_groups = Group.objects.get(name='Disallow')
            disAllow_groups.user_set.add(request.user)
            Addgroup = Group.objects.get(name='Allowed')
            Addgroup.user_set.remove(request.user)
        else:
            boolean = "false"
            disAllow_groups = Group.objects.get(name='Allowed')
            disAllow_groups.user_set.add(request.user)
            Addgroup = Group.objects.get(name='Disallow')
            Addgroup.user_set.remove(request.user)
        groupname = []
        for g in request.user.groups.all():
            groupname.append(g.name)
        allow = "Allowed"
        if allow in groupname:
            request.session['paid'] = "true"
        else:
            request.session['paid'] = "false"
            if request.user.username == "admin":
                request.session['paid'] = "true"
    if group:
        cases = case_type.split('|')
        lawyers_list = Lawyer.objects.filter(case__case_type__in=cases). \
            annotate(cases_count=Count('case')).order_by('-cases_count')

        for l, lawyer in enumerate(lawyers_list):
            lawyers_list[l].all_cases_count = lawyer.case_set.count()
            request.session['pagenumber'] = 0
        request.session['casetype'] = group
        page = request.GET.get('page')
        paginator = Paginator(lawyers_list, 10)
        lawyers_list = paginator.get_page(page)
        context = {'lawyers': lawyers_list, 'case_type': group, 'group': case_type, 'session': request.session['paid'],
                   'casetype': request.session['casetype'], 'expiredate': user_date}
    else:
        file_name = case_type.replace('\\', ' ').replace('/', ' ')
        file = open(settings.MEDIA_ROOT + '\\' + file_name + '.txt', encoding='utf-8', errors='ignore')
        lawyers_list = []
        fullname_lists = []
        for lawyer in file.readlines():
            lawyers_list.append(json.loads(lawyer))
        request.session['pagenumber'] = 0
        request.session['casetype'] = case_type
        page = request.GET.get('page')
        paginator = Paginator(lawyers_list, 10)
        lawyers_list = paginator.get_page(page)

        context = {'lawyers': lawyers_list, 'case_type': case_type, 'session': request.session['paid'],
                   'casetype': request.session['casetype'], 'expiredate': user_date}
    return render(request, 'lawyers_by_case.html', context)


def cases_types(request):
    user_date = datetime.now()
    user_date = formats.date_format(user_date, "SHORT_DATETIME_FORMAT")
    case_type = request.GET.get('case_type', '')
    group = request.GET.get('group', '')
    if case_type and group:
        cases_list = Case.objects.filter(case_type__icontains=case_type).values('case_type').exclude(case_type=''). \
            annotate(count=Count('case_type')).order_by('-case_type')
        context = {'case': case_type, 'cases': '|'.join([case['case_type'] for case in cases_list]), 'group': 'on',
                   'expiredate': user_date}
    elif case_type and not group:
        cases_list = Case.objects.filter(case_type__icontains=case_type).values('case_type').exclude(case_type=''). \
            annotate(count=Count('case_type')).order_by('case_type')
        context = {'cases': cases_list, 'expiredate': user_date}
    else:
        try:
            file = open(settings.MEDIA_ROOT + '/case_types.txt', encoding='utf-8', errors='ignore')
            cases_list = [{'case_type': case, 'session': request.session['paid']} for case in file.read().split(',')]
            file.close()
        except:
            cases_list = []
        context = {'cases': cases_list, 'expiredate': user_date}
    return render(request, 'case_types.html', context)


def thank_you(request):
    case_type = request.GET.get('Thank_you', '')
    request.session['paid'] = 'true'
    context = {'cases': request.session['paid'], 'image': settings.MEDIA_URL + 'payment_logo.png',
               'pageNumber': request.session['pagenumber'], 'casetype': request.session['casetype']}
    return render(request, 'Thank_you.html', context)


def cases_table_view(request, **kwargs):
    lawyer_bar_card = kwargs['bar_card']
    cases = Case.objects.filter(lawyer__bar_card=lawyer_bar_card)

    case_type = request.GET.get('case_type', '').replace('_', '&')
    group = request.GET.get('group', '')
    if case_type:
        if group:
            cases = cases.filter(case_type__in=case_type.split('|'))
        else:
            cases = cases.filter(case_type=case_type)

    representing = []
    disposition = []

    cases = sorted(cases, key=lambda a: datetime.strptime(a.date_filed, '%m/%d/%Y'), reverse=True)
    for case in cases:
        rep = ''
        try:
            parties = fromstring(case.parties)
            # get representing
            try:
                for id, party in enumerate(parties.xpath('//tbody/tr/td[3]')):
                    representative = etree.tostring(party, with_tail=False, method='html')
                    if str(representative).__contains__(case.lawyer.last_name):
                        rep = parties.xpath('//tbody/tr/td[2]')[id].xpath('text()')[0]
                        break
            except:
                pass
                # print('error occurred at', case.case_number)
        except:
            pass
        representing.append(rep)

        dis = ''
        try:
            case_events = fromstring(case.case_events)
            # get disposition
            try:
                event_types = [x for x in case_events.xpath('//tbody/tr/td[2]/text()')]
                event_types_dis = [x for x in case_events.xpath('//tbody/tr/td[3]/text()')]
                for id, event_type in enumerate(event_types):
                    if event_type == 'Memorandum opinion issued' or event_type == 'Opinion issued' or \
                            event_type == 'memorandum opinion issued' or event_type == 'opinion issued':
                        dis = event_types_dis[id]
                        break
                if not dis:
                    dis = [x for x in case_events.xpath('//tbody/tr/td[3]/text()') if x != u'\xa0'][0]
            except:
                pass
                # print('error occurred at', case.case_number)
        except:
            pass
        disposition.append(dis)

    context = {'cases': cases, 'representing': representing, 'disposition': disposition}
    return render(request, 'cases_table.html', context)


def cases_datas(request, **kwargs):
    case_id = int(kwargs['case_id'])
    case = Case.objects.filter(id=case_id)[0]
    case.case_events = case.case_events.replace('a href="SearchMedia.aspx?',
                                                'a href="http://www.search.txcourts.gov/SearchMedia.aspx?')
    context = {'case': case}
    return render(request, 'cases_subdatas.html', context)


def ajax_case(request):
    id = request.GET['data']
    category = id.split('_')[0]
    id = id.split('_')[1]
    case = Case.objects.filter(id=id)[0]
    result = ''
    if category == 'events':
        result = "events::" + "<br>" + case.case_events.replace('a href="SearchMedia.aspx?',
                                                                'a href="http://www.search.txcourts.gov/SearchMedia.aspx?')
    elif category == 'trial':
        result = "trial::" + "<br>" + case.trial_court_information
    elif category == 'parties':
        result = "parties::" + "<br>" + case.parties
    elif category == 'calendars':
        result = "calendars::" + "<br>" + case.calendars
    elif category == 'briefs':
        result = "briefs::" + "<br>" + case.appellate_briefs.replace('a href="SearchMedia.aspx?',
                                                                     'a href="http://www.search.txcourts.gov/SearchMedia.aspx?')
    return HttpResponse(result)


class LawyerView(DetailView):
    model = Lawyer
    template_name = 'lawyer.html'
    slug_url_kwarg = 'bar_card'
    context_object_name = 'lawyer'

    def get_context_data(self, **kwargs):
        context = super(LawyerView, self).get_context_data(**kwargs)
        cases = Case.objects.filter(lawyer__bar_card=self.kwargs['pk'])
        context['cases'] = cases

        # To show coming from case at the top
        coming_from_case_type = self.request.GET.get('case_type', '').replace('_', '&')
        group = self.request.GET.get('group', '')

        filtered_cases = cases.values('case_type').exclude(case_type=''). \
            annotate(count_type=Count('case_type')).order_by('-count_type')

        if coming_from_case_type:
            if group:
                result_cases = []
                case_list = [a.replace('_', '&') for a in coming_from_case_type.split('|')]
                case_types = [f['case_type'] for f in filtered_cases]
                for item in case_list:
                    try:
                        i = case_types.index(item)
                    except:
                        continue
                    a = filtered_cases[i]
                    a['highlight'] = 'on'
                    result_cases.append(a)
                result_cases = sorted(result_cases, key=lambda a: a['count_type'], reverse=True)
                for item in filtered_cases:
                    if item['case_type'] not in case_list:
                        result_cases.append(item)

                context['filtered_cases'] = result_cases
            else:
                case_types = [f['case_type'] for f in filtered_cases]
                i = case_types.index(coming_from_case_type)
                filtered_cases = filtered_cases[i:i + 1] + filtered_cases[:i] + filtered_cases[i + 1:]

                context['filtered_cases'] = filtered_cases
        else:
            context['filtered_cases'] = filtered_cases

        return context


###### Functionalities done by baig052@gmail.com ###

def admin_login(request):
    if request.method == "POST":
        username = request.POST['email']
        user = auth.authenticate(username=request.POST['email'], password=request.POST['password'])
        if user is not None:
            auth.login(request, user)
            return redirect('admin_dashboard')
        else:
            return render(request, 'admin_login.html', {'error': 'Email/Password is Incorrect.'})
    else:
        context = {}
        return render(request, 'admin_login.html', context)


@login_required(login_url="admin_login")
def admin_dashboard(request):
    datatable = AssignLinks.objects
    if request.method == 'POST':
        if request.POST.get('user_id') and request.POST.get('weblinkallowed') and request.POST.get('numberoftime'):
            datatable = AssignLinks()
            datatable.email = request.POST['user_id']
            datatable.link_id = request.POST['weblinkallowed']
            datatable.allowed = int(request.POST['numberoftime'])
            datatable.left = request.POST['numberoftime']
            datatable.date = datetime.now()
            datatable.save()
            return redirect('admin_dashboard')

        if request.POST.get('delete'):
            AssignLinks.objects.filter(id__in=request.POST.getlist('items')).delete()

    return render(request, 'admin_dashboard.html', {'datatable': datatable})


def admin_logout(request):
    if request.method == 'POST':
        auth.logout(request)
        return redirect('admin_login')


@login_required(login_url="admin_login")
def admin_time(request, datatable_id):
    timeinfo = SeenDetails.objects.filter(assign_links_id=datatable_id)
    assign_link = AssignLinks.objects.filter(id=datatable_id)
    if not assign_link:
        return HttpResponse("No Id Found")
    return render(request, 'admin_time.html', {'timeinfos': timeinfo,
                                               'assignlink': assign_link[0]})

###### END Functionalities done by baig052@gmail.com ###
